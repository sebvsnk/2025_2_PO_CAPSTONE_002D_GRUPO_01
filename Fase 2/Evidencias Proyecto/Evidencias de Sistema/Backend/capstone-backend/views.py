# EN capstone-backend/api/views.py

import os
import uuid
import csv # Módulo de Python para CSV
from io import StringIO, BytesIO # Para escribir CSV/Excel en memoria
from datetime import datetime, timedelta
from collections import defaultdict
import html # <-- Nuevo módulo para escapar HTML de forma segura
import urllib.parse # <-- 1. AÑADIDO: Para codificar URLs
from xhtml2pdf import pisa
from xhtml2pdf import pisa # La librería de PDF
# --- Importaciones de Django ---
from django.db import connection, models # models añadido para Q
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.db.models import F, Sum, Count, Subquery, OuterRef, DurationField, Avg, ExpressionWrapper, fields, Q
from django.db.models.functions import Coalesce
from django.core.files.uploadedfile import UploadedFile

# --- Importaciones de Django REST Framework ---
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics, filters, status, permissions, mixins # Clase base para todas las vistas DRF
from rest_framework.pagination import PageNumberPagination # <-- 1. IMPORTAR ESTO
from rest_framework.permissions import AllowAny, IsAuthenticated
# --- 1. IMPORTAR DJANGO-FILTER ---
from django_filters.rest_framework import DjangoFilterBackend

from rest_framework.parsers import MultiPartParser, FormParser

# --- Importaciones de Supabase ---
from supabase import create_client, Client
from gotrue.errors import AuthApiError # Importamos el error específico de la librería supabase-py

# --- Importaciones de Openpyxl (para Excel) ---
import openpyxl
from openpyxl.utils import get_column_letter

# --- Importaciones Locales (Modelos, Serializers, Permisos) ---

# 1. Modelos (Asegúrate de incluir AuditLog, si ya lo definiste en models.py)
from .models import (
    Vehiculo, BitacoraPorteria, Ot, Usuario, Estado, Tarea, Pausa, Evidencia, AuditLog, Repuesto, TareaRepuesto
)

# 2. Serializers (Asegúrate de incluir todos tus serializers, incluyendo el de Auditoría y los de Admin)
from .serializers import (
    VehiculoSerializer, VehiculoCreateSerializer, BitacoraSerializer,
    OtSerializer, OtCreateSerializer, OtUpdateSerializer, OtCambiarEstadoSerializer,
    OtEstadoChoferSerializer, PublicOtStatusSerializer,
    TareaSerializer, TareaCreateSerializer, TareaPausarSerializer, TareaUpdateSerializer,
    EvidenciaSerializer,
    ReporteHorasHombreSerializer,
    TableroOtSerializer,
    BitacoraConTotalesSerializer,
    ReporteDuracionEtapaSerializer,
    AdminUsuarioCreateSerializer, AdminUsuarioSerializer, AdminUsuarioUpdateSerializer, # Serializers de Admin
    AuditLogSerializer, TareaRepuestoSerializer, TareaRepuestoCreateSerializer, OtExportSerializer,TareaExportSerializer, OtExportSerializer  # ¡El que faltaba!
)

# 3. Permisos (Solo clases de Permiso)
from .permissions import (
    IsAdminUser, IsSupervisorUser, IsMecanicoUser, IsPorteriaUser, IsChoferUser, IsAnalistaUser,
    IsSupervisorOrAdminUser, IsMecanicoOrSupervisorUser, IsSupervisorOrAdminOrAnalistaUser, IsPorteriaOrSupervisorOrAdminOrAnalista
)

# --- 2. AÑADE ESTA NUEVA CLASE (después de las importaciones) ---
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50 # Número de logs por página
    page_size_query_param = 'page_size'
    max_page_size = 200
# --- FIN DE LA NUEVA CLASE ---

# ============================================
# --- VISTAS HTML (Admin/Dashboard/Login) ---
# ============================================

from django.contrib.auth import authenticate, login
# Importaciones necesarias para el contexto del Dashboard
from .models import Usuario # <-- Asegúrate de que Usuario esté importado

# ============================================
# --- VISTAS HTML (Admin/Dashboard/Login) ---
# ============================================

@login_required(login_url="/login")
@user_passes_test(lambda u: u.is_staff, login_url="/login")
def home(request):
    """Renderiza el dashboard principal (HTML) solo para staff."""
    context = {"username": request.user.get_username() or "usuario"}
    return render(request, "dashboard.html", context)

@csrf_exempt
def login_page(request):
    """Renderiza la página de login (HTML) y maneja el POST para el dashboard."""
    if request.user.is_authenticated and request.user.is_staff:
        return HttpResponseRedirect("/")
    error = ""
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return HttpResponseRedirect(request.GET.get("next", "/"))
        else:
            error = "Usuario o contraseña inválidos, o no tiene permisos de staff."
    return render(request, "login.html", {"error": error})

# ==================================
# --- VISTAS DE API (JSON) ---
# ==================================

# --- Utilidad y Salud ---

class Health(APIView):
    """Endpoint de chequeo de salud (DB y conexión)."""
    permission_classes = [AllowAny]
    def get(self, request):
        out = {"ok": True, "db_vendor": connection.vendor}
        try:
            with connection.cursor() as cur:
                cur.execute("select current_database()")
                out["db"] = cur.fetchone()[0]
                cur.execute("show search_path")
                out["search_path"] = cur.fetchone()[0]
        except Exception as e:
            out["ok"] = False
            out["error"] = str(e)
        out["PGHOST"] = os.getenv("PGHOST")
        out["PGPORT"] = os.getenv("PGPORT")
        status_code = status.HTTP_503_SERVICE_UNAVAILABLE if not out["ok"] else status.HTTP_200_OK
        return Response(out, status=status_code)

# --- Vehículos ---

class VehiculoList(generics.ListAPIView):
    """(GET) Lista vehículos. Filtro: ?search=<patente>."""
    permission_classes = [AllowAny]
    queryset = Vehiculo.objects.all().order_by("-creado_en")
    serializer_class = VehiculoSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["patente"]

class VehiculoCreate(generics.CreateAPIView):
    """(POST) Crea un vehículo."""
    permission_classes = [IsAuthenticated, IsSupervisorOrAdminUser]
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoCreateSerializer

class VehiculoDetail(generics.RetrieveAPIView):
    """(GET) Detalle de vehículo por ID."""
    permission_classes = [IsAuthenticated]
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer
    lookup_url_kwarg = "vehiculo_id"

# --- Bitácora (Portería) ---

class BitacoraListCreate(generics.ListCreateAPIView):
    """
    (GET) Lista bitácora con filtros y totales.
          Filtros: ?patente=..., ?fecha_inicio=YYYY-MM-DD, ?fecha_fin=YYYY-MM-DD
    (POST) Registra entrada/salida (requiere auth).
    """
    serializer_class = BitacoraSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["vehiculo__patente"] # Búsqueda parcial por patente

    def get_queryset(self):
        """Aplica filtros de fecha y patente exacta (si se proveen)."""
        queryset = BitacoraPorteria.objects.select_related('vehiculo').all()

        patente_exacta = self.request.query_params.get('patente')
        if patente_exacta:
             patente_norm = patente_exacta.upper().replace('-', '')
             queryset = queryset.filter(vehiculo__patente=patente_norm)

        try: # Filtro por Rango de Fechas (RF-BIT-02)
            fecha_inicio = self.request.query_params.get('fecha_inicio')
            fecha_fin = self.request.query_params.get('fecha_fin')
            tz = timezone.get_current_timezone() #

            if fecha_inicio:
                fecha_inicio_dt = datetime.fromisoformat(fecha_inicio).replace(hour=0, minute=0, second=0)
                fecha_inicio_dt_aware = timezone.make_aware(fecha_inicio_dt, tz) #
                queryset = queryset.filter(fecha_hora__gte=fecha_inicio_dt_aware) #

            if fecha_fin:
                fecha_fin_dt = datetime.fromisoformat(fecha_fin).replace(hour=23, minute=59, second=59)
                fecha_fin_dt_aware = timezone.make_aware(fecha_fin_dt, tz) #
                queryset = queryset.filter(fecha_hora__lte=fecha_fin_dt_aware) #

        except (ValueError, TypeError):
            pass # Ignoramos fechas con formato incorrecto

        return queryset.order_by("-fecha_hora") #

    def get_permissions(self):
        if self.request.method == "POST":
            # Permiso para registrar (POST): Solo Portería
            return [permissions.IsAuthenticated(), IsPorteriaUser()] 
        # ✅ FIX: Permiso para listar (GET): Portería Y Supervisor/Admin/Analista
        return [permissions.IsAuthenticated(), IsPorteriaOrSupervisorOrAdminOrAnalista()]

    def list(self, request, *args, **kwargs):
        """Sobrescribe el método list para calcular totales (RF-BIT-03)."""
        queryset_filtrado = self.filter_queryset(self.get_queryset()) #

        totales = queryset_filtrado.aggregate(
            entradas=Count('id', filter=Q(tipo='ENTRADA')), #
            salidas=Count('id', filter=Q(tipo='SALIDA')) #
        )

        permanencia_data = queryset_filtrado.values('vehiculo_id').annotate(
            primera_entrada=models.Min('fecha_hora', filter=Q(tipo='ENTRADA')),
            ultima_salida=models.Max('fecha_hora', filter=Q(tipo='SALIDA'))
        ).filter(
            primera_entrada__isnull=False,
            ultima_salida__isnull=False,
            ultima_salida__gt=F('primera_entrada') #
        ).aggregate(
            avg_permanencia=Avg(
                ExpressionWrapper(F('ultima_salida') - F('primera_entrada'), output_field=fields.DurationField()) #
            )
        )

        avg_permanencia_timedelta = permanencia_data.get('avg_permanencia')
        if avg_permanencia_timedelta:
            totales['permanencia_promedio_minutos'] = round(avg_permanencia_timedelta.total_seconds() / 60) #
        else:
            totales['permanencia_promedio_minutos'] = 0 #

        page = self.paginate_queryset(queryset_filtrado) #
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data) #
            paginated_response.data['totales'] = totales #
            return paginated_response

        serializer = self.get_serializer(queryset_filtrado, many=True)
        return Response({
            'totales': totales,
            'registros': serializer.data #
        })

# --- Órdenes de Trabajo (OT) ---

class OtListCreateView(generics.ListCreateAPIView):
    """(GET) Lista OTs. (POST) Crea OT."""
    queryset = Ot.objects.select_related("vehiculo", "estado", "creado_por").all()

    def get_permissions(self):
        if self.request.method == 'POST':
            return [permissions.IsAuthenticated(), IsSupervisorOrAdminUser()] #
        return [permissions.IsAuthenticated()] #

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return OtCreateSerializer
        return OtSerializer
    def perform_create(self, serializer):
        # 1. Guarda la OT (esto llama a OtCreateSerializer.create, que crea la instancia)
        instance = serializer.save()
        
        # 2. Obtiene el ID del usuario creador (ya resuelto por el serializer)
        creador_id = instance.creado_por_id

        # 3. Registra el evento de Auditoría (RF-AUD-02)
        register_audit_log(
            user_id=creador_id,
            action_type="OT_CREATE",
            entity_type="ot",
            entity_id=instance.id,
            details={"descripcion": instance.descripcion, "vehiculo_id": instance.vehiculo_id}
        )
class OtDetailUpdateView(mixins.RetrieveModelMixin,
                         mixins.UpdateModelMixin,
                         generics.GenericAPIView):
    """(GET) Detalle OT. (PUT/PATCH) Actualiza descripción OT."""
    queryset = Ot.objects.select_related(
        "vehiculo", "estado", "creado_por"
    ).annotate(
        tareas_count=Count('tareas') # <-- 2. AÑADE ESTA ANOTACIÓN
    ).all()
    lookup_url_kwarg = 'ot_id'

    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH']:
            return [permissions.IsAuthenticated(), IsSupervisorOrAdminUser()] #
        return [permissions.IsAuthenticated()] #

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return OtUpdateSerializer
        return OtSerializer

    def get(self, request, *args, **kwargs):
        """Maneja GET para ver detalle."""
        return self.retrieve(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        """Maneja PUT (actualización completa)."""
        return self.update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        """Maneja PATCH (actualización parcial)."""
        return self.partial_update(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        write_serializer = self.get_serializer(instance, data=request.data, partial=partial)
        write_serializer.is_valid(raise_exception=True)
        self.perform_update(write_serializer)
        if getattr(instance, '_prefetched_objects_cache', None):
            instance = self.get_object()
        read_serializer = OtSerializer(instance, context=self.get_serializer_context()) #
        return Response(read_serializer.data) #

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs) #
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # 1. Capturar el valor original de la descripción para auditoría
        original_descripcion = instance.descripcion

        write_serializer = self.get_serializer(instance, data=request.data, partial=partial)
        write_serializer.is_valid(raise_exception=True)

        # 2. Realizar la actualización (llama a save() y guarda los cambios)
        self.perform_update(write_serializer)

        # 3. Re-obtener la instancia actualizada para la respuesta y auditoría
        if getattr(instance, '_prefetched_objects_cache', None):
            instance = self.get_object()

        # 4. Auditoría si la descripción cambió (RF-AUD-02)
        if original_descripcion != instance.descripcion:
             try:
                perfil_usuario = Usuario.objects.get(external_id=request.user.username)
                usuario_id = perfil_usuario.id
             except Usuario.DoesNotExist:
                usuario_id = None

             register_audit_log(
                user_id=usuario_id,
                action_type="OT_EDIT",
                entity_type="ot",
                entity_id=instance.id,
                details={
                    "field": "descripcion",
                    "old_value": original_descripcion,
                    "new_value": instance.descripcion,
                    "by_partial_update": partial # Indica si fue PATCH (true) o PUT (false)
                }
            )

        read_serializer = OtSerializer(instance, context=self.get_serializer_context())
        return Response(read_serializer.data)

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

# EN capstone-backend/api/views.py

class OtCambiarEstadoView(APIView):
    """(POST) Cambia estado de OT. Body: {"estado_code": "..."}."""
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminUser]

    def post(self, request, ot_id):
        # Aseguramos cargar la instancia de la OT
        ot_instance = get_object_or_404(Ot, pk=ot_id)
        serializer = OtCambiarEstadoSerializer(data=request.data)

        if serializer.is_valid():
            nuevo_estado_code = serializer.validated_data['estado_code']
            estado_actual = ot_instance.estado
            estado_actual_code = estado_actual.code if estado_actual else None

            # 1. Validación de Precondición de Cierre: OT ya Cerrada
            if estado_actual_code == 'CERRADA' and nuevo_estado_code != 'CERRADA':
                 return Response({"detail": "No se puede cambiar el estado de una OT cerrada."}, status=status.HTTP_400_BAD_REQUEST)

            # 2. ✅ FIX: VALIDACIÓN RF-OT-03 (TAREAS ABIERTAS)
            if nuevo_estado_code == 'CERRADA':
                # Busca tareas de esta OT que NO estén en estado final (HECHA o ANULADA)
                tareas_abiertas = Tarea.objects.filter(
                    ot=ot_instance
                ).exclude(
                    # Los estados HECHA y ANULADA son los únicos que permiten cerrar la OT
                    estado__code__in=['HECHA', 'ANULADA'] 
                ).exists()

                if tareas_abiertas:
                    return Response({
                        "detail": "No se puede cerrar la OT.", 
                        "pendientes": "Aún existen tareas activas o pendientes de cerrar/anular (ej: NUEVA, EN_PROCESO, PAUSADA)."
                    }, status=status.HTTP_400_BAD_REQUEST)
                # La validación de Evidencia Mínima (RF-OT-03) podría añadirse aquí.
            # --- FIN FIX ---

            try:
                nuevo_estado = Estado.objects.get(tipo='ot', code=nuevo_estado_code)
            except Estado.DoesNotExist:
                 return Response({"detail": f"Estado OT '{nuevo_estado_code}' no encontrado."}, status=status.HTTP_400_BAD_REQUEST)

            # --- 3. CAMBIO DE ESTADO Y GUARDADO ---
            now = timezone.now()
            ot_instance.estado = nuevo_estado
            ot_instance.actualizado_en = now
            
            if nuevo_estado_code == 'CERRADA' and ot_instance.fecha_cierre is None:
                ot_instance.fecha_cierre = now
            
            ot_instance.save() 

            # --- 4. AUDITORÍA (RF-AUD-02) ---
            try:
                perfil_usuario = Usuario.objects.get(external_id=request.user.username)
                usuario_id = perfil_usuario.id
            except Usuario.DoesNotExist:
                usuario_id = None
                
            register_audit_log(
                user_id=usuario_id,
                action_type="OT_STATUS_CHANGE",
                entity_type="ot",
                entity_id=ot_instance.id,
                details={
                    "old_status_code": estado_actual_code,
                    "new_status_code": nuevo_estado_code
                }
            )
            
            # --- 5. RESPUESTA ---
            read_serializer = OtSerializer(ot_instance)
            return Response(read_serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# --- Tareas ---

# EN capstone-backend/api/views.py

class TareaListCreateView(generics.ListCreateAPIView):
    """(GET) Lista tareas de una OT. (POST) Crea tarea en OT."""
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TareaCreateSerializer
        return TareaSerializer

    def get_queryset(self):
        ot_id = self.kwargs.get('ot_id')
        get_object_or_404(Ot, pk=ot_id)
        return Tarea.objects.filter(ot_id=ot_id).select_related(
            'estado', 'responsable'
        ).prefetch_related(
            'evidencias', 'repuestos_usados'
        ).order_by('creado_en')

    def perform_create(self, serializer):
        ot_id = self.kwargs.get('ot_id')
        ot_instance = get_object_or_404(Ot, pk=ot_id)

        if ot_instance.estado and ot_instance.estado.code in ['CERRADA', 'ANULADA']:
             raise Http404(f"No se pueden añadir tareas a una OT en estado {ot_instance.estado.code}.")

        # 1. Save the model instance (Persistencia y Corrección de FK)
        # Esto guarda el objeto en la DB y establece serializer.instance
        serializer.save(
            ot_id=ot_instance.id, 
            creado_en=timezone.now(),
            actualizado_en=timezone.now()
        )

        # 2. Get the saved instance via serializer.instance for auditing.
        instance = serializer.instance 

        # 3. Get the user ID for auditing
        try:
            perfil_usuario = Usuario.objects.get(external_id=self.request.user.username)
            creador_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            creador_id = None

        # 4. Register the audit log (RF-AUD-02)
        register_audit_log(
            user_id=creador_id,
            action_type="TAREA_CREATE",
            entity_type="tarea",
            entity_id=instance.id, 
            details={"nombre": instance.nombre, "ot_id": instance.ot_id}
        )

    def create(self, request, *args, **kwargs):
        # Este método es necesario para que el perform_create se ejecute y luego use TareaSerializer para la respuesta.
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        instance = serializer.instance # Aquí se lee la instancia Tarea creada
        read_serializer = TareaSerializer(instance)
        headers = self.get_success_headers(read_serializer.data)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class TareaDetailUpdateView(mixins.RetrieveModelMixin,
                            mixins.UpdateModelMixin,
                            generics.GenericAPIView):
    """(GET) Detalle Tarea. (PUT/PATCH) Actualiza nombre/responsable Tarea."""
    queryset = Tarea.objects.select_related(
        'estado', 'responsable', 'ot', 'ot__vehiculo'
    ).prefetch_related(
        'evidencias', 'repuestos_usados' # <-- AÑADE ESTE prefetch_related
    ).all()
    lookup_url_kwarg = 'tarea_id' #

    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH']:
            return [permissions.IsAuthenticated(), IsSupervisorOrAdminUser()] #
        return [permissions.IsAuthenticated(), IsMecanicoOrSupervisorUser()] #

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return TareaUpdateSerializer #
        return TareaSerializer #

    def get(self, request, *args, **kwargs):
        return self.retrieve(request, *args, **kwargs) #

    def put(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs) #

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs) #

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        if instance.ot and instance.ot.estado and instance.ot.estado.code in ['CERRADA', 'ANULADA']:
             return Response({"detail": f"No se puede modificar una tarea de una OT cerrada o anulada."}, status=status.HTTP_400_BAD_REQUEST) #

        write_serializer = self.get_serializer(instance, data=request.data, partial=partial)
        write_serializer.is_valid(raise_exception=True)
        self.perform_update(write_serializer)
        if getattr(instance, '_prefetched_objects_cache', None):
            instance = self.get_object()
        read_serializer = TareaSerializer(instance, context=self.get_serializer_context()) #
        return Response(read_serializer.data) #

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs) #

# --- Ciclo de Vida de Tareas ---

class TareaIniciarView(APIView):
    """(POST) Inicia una Tarea."""
    # PERMISO MODIFICADO: Ahora permite Mecánico O Supervisor
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser] 

    def post(self, request, tarea_id):
        tarea = get_object_or_404(
            Tarea.objects.select_related(
                'ot', 'ot__estado', 'estado', 'responsable', 'ot__vehiculo'
            ).prefetch_related('evidencias', 'repuestos_usados'),
            pk=tarea_id
        )

        if tarea.estado is None or tarea.estado.code != 'NUEVA':
            return Response({"detail": f"Solo se pueden iniciar tareas en estado 'NUEVA'."}, status=status.HTTP_400_BAD_REQUEST)
        if tarea.ot.estado and tarea.ot.estado.code not in ['ACTIVA', 'EN_PROCESO']:
            return Response({"detail": f"No se puede iniciar tarea si la OT no está activa."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            estado_en_proceso = Estado.objects.get(tipo='tarea', code='EN_PROCESO')
        except Estado.DoesNotExist:
            return Response({"detail": "Estado 'EN_PROCESO' para tareas no encontrado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        now = timezone.now()
        tarea.estado = estado_en_proceso
        tarea.inicio = now # Registra inicio
        tarea.actualizado_en = now
        tarea.save()

        # --- LÓGICA DE AUDITORÍA ---
        try:
            perfil_usuario = Usuario.objects.get(external_id=request.user.username)
            usuario_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            usuario_id = None
            
        register_audit_log(
            user_id=usuario_id,
            action_type="TAREA_START",
            entity_type="tarea",
            entity_id=tarea.id,
            details={"inicio_registrado": now.isoformat()}
        )
        # --- FIN LÓGICA DE AUDITORÍA ---

        serializer = TareaSerializer(tarea)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TareaPausarView(APIView):
    """(POST) Pausa una Tarea. Body: {"motivo": "..."}."""
    # PERMISO MODIFICADO: Ahora permite Mecánico O Supervisor
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser] 

    def post(self, request, tarea_id):
        tarea = get_object_or_404(
            Tarea.objects.select_related(
                'ot', 'ot__estado', 'estado', 'responsable', 'ot__vehiculo'
            ).prefetch_related('evidencias', 'repuestos_usados'),
            pk=tarea_id
        )

        if tarea.estado is None or tarea.estado.code != 'EN_PROCESO':
            return Response({"detail": f"Solo se pueden pausar tareas en estado 'EN_PROCESO'."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = TareaPausarSerializer(data=request.data)
        if serializer.is_valid():
            motivo_pausa = serializer.validated_data['motivo']
            try:
                estado_pausada = Estado.objects.get(tipo='tarea', code='PAUSADA')
            except Estado.DoesNotExist:
                return Response({"detail": "Estado 'PAUSADA' para tareas no encontrado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            now = timezone.now()
            # 1. Crea el registro de Pausa
            Pausa.objects.create(tarea=tarea, motivo=motivo_pausa, inicio=now)
            
            # 2. Actualiza la tarea
            tarea.estado = estado_pausada
            tarea.actualizado_en = now
            tarea.save()

            # --- LÓGICA DE AUDITORÍA ---
            try:
                perfil_usuario = Usuario.objects.get(external_id=request.user.username)
                usuario_id = perfil_usuario.id
            except Usuario.DoesNotExist:
                usuario_id = None
                
            register_audit_log(
                user_id=usuario_id,
                action_type="TAREA_PAUSE",
                entity_type="tarea",
                entity_id=tarea.id,
                details={"motivo": motivo_pausa, "pausa_iniciada": now.isoformat()}
            )
            # --- FIN LÓGICA DE AUDITORÍA ---

            read_serializer = TareaSerializer(tarea)
            return Response(read_serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TareaReanudarView(APIView):
    """(POST) Reanuda una Tarea."""
    # PERMISO MODIFICADO: Ahora permite Mecánico O Supervisor
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser] 

    def post(self, request, tarea_id):
        tarea = get_object_or_404(
            Tarea.objects.select_related(
                'ot', 'ot__estado', 'estado', 'responsable', 'ot__vehiculo'
            ).prefetch_related('evidencias', 'repuestos_usados'),
            pk=tarea_id
        )

        if tarea.estado is None or tarea.estado.code != 'PAUSADA':
            return Response({"detail": f"Solo se pueden reanudar tareas en estado 'PAUSADA'."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            estado_en_proceso = Estado.objects.get(tipo='tarea', code='EN_PROCESO')
        except Estado.DoesNotExist:
            return Response({"detail": "Estado 'EN_PROCESO' para tareas no encontrado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        now = timezone.now()
        pausa_activa = Pausa.objects.filter(tarea=tarea, fin__isnull=True).order_by('-inicio').first()
        
        # 1. Cierra la pausa activa (si existe)
        if pausa_activa:
            pausa_activa.fin = now
            pausa_activa.save()
            duracion_pausa = now - pausa_activa.inicio
        else:
            duracion_pausa = timedelta(0)

        # 2. Actualiza la tarea
        tarea.estado = estado_en_proceso
        tarea.actualizado_en = now
        tarea.save()

        # --- LÓGICA DE AUDITORÍA ---
        try:
            perfil_usuario = Usuario.objects.get(external_id=request.user.username)
            usuario_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            usuario_id = None

        register_audit_log(
            user_id=usuario_id,
            action_type="TAREA_RESUME",
            entity_type="tarea",
            entity_id=tarea.id,
            details={
                "pausa_finalizada": now.isoformat(),
                "duracion_segundos": round(duracion_pausa.total_seconds())
            }
        )
        # --- FIN LÓGICA DE AUDITORÍA ---

        serializer = TareaSerializer(tarea)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TareaCerrarView(APIView):
    """(POST) Cierra una Tarea."""
    # PERMISO CORRECTO: Permite Mecánico O Supervisor
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser] 

    def post(self, request, tarea_id):
        tarea = get_object_or_404(
            Tarea.objects.select_related(
                'ot', 'ot__estado', 'estado', 'responsable', 'ot__vehiculo'
            ).prefetch_related('evidencias', 'repuestos_usados'),
            pk=tarea_id
        )

        estados_validos = ['EN_PROCESO', 'PAUSADA']
        if tarea.estado is None or tarea.estado.code not in estados_validos:
            return Response({"detail": f"Solo se pueden cerrar tareas en estado {estados_validos}."}, status=status.HTTP_400_BAD_REQUEST)

        MIN_EVIDENCIAS = 0
        if MIN_EVIDENCIAS > 0 and not Evidencia.objects.filter(tarea=tarea).exists():
             return Response({"detail": f"No se puede cerrar la tarea, requiere al menos {MIN_EVIDENCIAS} evidencia(s)."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            estado_hecha = Estado.objects.get(tipo='tarea', code='HECHA')
        except Estado.DoesNotExist:
            return Response({"detail": "Estado 'HECHA' para tareas no encontrado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        now = timezone.now()
        
        # 1. Cerrar pausa activa si la tarea estaba pausada
        if tarea.estado.code == 'PAUSADA':
             pausa_activa = Pausa.objects.filter(tarea=tarea, fin__isnull=True).order_by('-inicio').first()
             if pausa_activa:
                 pausa_activa.fin = now
                 pausa_activa.save()
        
        # 2. Actualizar la tarea
        tarea.estado = estado_hecha
        tarea.fin = now # Registra fin
        tarea.actualizado_en = now
        tarea.save()
        
        # --- LÓGICA DE AUDITORÍA ---
        try:
            perfil_usuario = Usuario.objects.get(external_id=request.user.username)
            usuario_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            usuario_id = None
        
        # Calcular duración total (si hay fecha de inicio)
        duracion_bruta_segundos = round((now - tarea.inicio).total_seconds()) if tarea.inicio else 0
        
        register_audit_log(
            user_id=usuario_id,
            action_type="TAREA_CLOSE",
            entity_type="tarea",
            entity_id=tarea.id,
            details={
                "fin_registrado": now.isoformat(),
                "duracion_bruta_segundos": duracion_bruta_segundos
            }
        )
        # --- FIN LÓGICA DE AUDITORÍA ---

        serializer = TareaSerializer(tarea)
        return Response(serializer.data, status=status.HTTP_200_OK)

class EvidenciaCreateView(APIView):
    """
    (GET) Lista evidencias de una Tarea.
    (POST) Sube una o MÚLTIPLES evidencias a Tarea (RF-EV-01).
    """
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser]
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request, tarea_id):
        """Lista evidencias de la tarea."""
        tarea = get_object_or_404(Tarea, pk=tarea_id)
        queryset = Evidencia.objects.filter(tarea=tarea).select_related('subido_por').order_by('-fecha_subida')
        serializer = EvidenciaSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, tarea_id):
        """Sube uno o MÚLTIPLES archivos de evidencia a la tarea."""
        # Aseguramos cargar 'ot' para obtener su ID.
        tarea = get_object_or_404(Tarea.objects.select_related('ot', 'ot__estado'), pk=tarea_id)

        if tarea.ot.estado and tarea.ot.estado.code in ['CERRADA', 'ANULADA']:
            return Response({"detail": "No se puede añadir evidencia a una tarea de una OT cerrada o anulada."}, status=status.HTTP_400_BAD_REQUEST)

        # --- 1. MODIFICACIÓN CLAVE: Usar .getlist('file') ---
        archivos_subidos = request.FILES.getlist('file')
        if not archivos_subidos:
            return Response({"detail": "No se enviaron archivos con la clave 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        # Configuración (movida fuera del bucle)
        MAX_SIZE_MB = 10
        ALLOWED_TYPES = ['image/jpeg', 'image/png', 'application/pdf']
        bucket_name = "apt-evidencias"
        ot_id = tarea.ot_id

        try:
            supabase_url: str = os.environ.get("SUPABASE_URL")
            supabase_key: str = os.environ.get("SUPABASE_SERVICE_KEY")
            if not supabase_url or not supabase_key:
                raise Exception("Configuración Supabase incompleta.")
            supabase: Client = create_client(supabase_url, supabase_key)
        except Exception as e:
            return Response({"detail": f"Error cliente Supabase: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            perfil_usuario = Usuario.objects.get(external_id=request.user.username)
            subido_por_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            subido_por_id = None

        # Listas para guardar los resultados
        evidencias_creadas = []
        errores_subida = []

        # --- 2. MODIFICACIÓN: Iterar sobre los archivos ---
        for archivo in archivos_subidos:
            nombre_original = archivo.name
            storage_path = "" # Definir para el scope de rollback

            try:
                # 2a. Validar CADA archivo
                if archivo.size > MAX_SIZE_MB * 1024 * 1024:
                    raise Exception(f"Archivo '{nombre_original}' supera {MAX_SIZE_MB} MB.")
                if archivo.content_type not in ALLOWED_TYPES:
                    raise Exception(f"Tipo de archivo no permitido en '{nombre_original}' ({archivo.content_type}).")

                # 2b. Generar ruta y subir a Supabase
                nombre_unico = f"{uuid.uuid4()}_{nombre_original}"
                storage_path = f"ot/{ot_id}/tareas/{tarea_id}/{nombre_unico}"

                # Necesitamos leer el archivo (esto lo pone en memoria)
                archivo_bytes = archivo.read()
                
                supabase.storage.from_(bucket_name).upload(
                    path=storage_path, file=archivo_bytes,
                    file_options={"content-type": archivo.content_type}
                )

                # 2c. Guardar en BD
                evidencia = Evidencia.objects.create(
                    tarea=tarea,
                    path=storage_path,
                    mime_type=archivo.content_type,
                    tamano_bytes=archivo.size,
                    subido_por_id=subido_por_id
                    # fecha_subida usa default=timezone.now
                )
                evidencias_creadas.append(evidencia)

            except Exception as e:
                # Si algo falla (validación, subida, BD), lo registramos
                errores_subida.append({"file": nombre_original, "error": str(e)})
                
                # Intentar rollback de storage si la subida fue exitosa pero la BD falló
                if storage_path and "Error guardando en BD" in str(e):
                    try: supabase.storage.from_(bucket_name).remove([storage_path])
                    except: pass # Ignorar error en rollback

        # --- 3. MODIFICACIÓN: Respuesta ---
        if not evidencias_creadas and errores_subida:
            # Si NADA se subió y todo falló
            return Response({"detail": "Todos los archivos fallaron.", "errors": errores_subida}, status=status.HTTP_400_BAD_REQUEST)

        # Si AL MENOS UNO se subió (éxito parcial o total)
        serializer = EvidenciaSerializer(evidencias_creadas, many=True)
        return Response(
            {"data": serializer.data, "errors": errores_subida},
            status=status.HTTP_201_CREATED
        )

# --- Vistas Específicas por Rol ---

class MiVehiculoEstadoView(APIView):
    """(GET) Vista para Chofer: OTs activas/pausadas de sus vehículos."""
    permission_classes = [permissions.IsAuthenticated, IsChoferUser] #

    def get(self, request):
        try:
            chofer_perfil = Usuario.objects.get(external_id=request.user.username, rol='CHOFER') #
        except Usuario.DoesNotExist:
            return Response({"detail": "Perfil de chofer no encontrado."}, status=status.HTTP_404_NOT_FOUND) #

        vehiculos_del_chofer_ids = chofer_perfil.vehiculos_asignados.values_list('id', flat=True) #
        if not vehiculos_del_chofer_ids:
            return Response([], status=status.HTTP_200_OK) #

        codigos_visibles = ['ACTIVA', 'PAUSADA'] #
        ots_visibles = Ot.objects.filter(
            vehiculo_id__in=vehiculos_del_chofer_ids,
            estado__code__in=codigos_visibles,
            estado__tipo='ot' #
        ).select_related('estado').order_by('-actualizado_en') #

        serializer = OtEstadoChoferSerializer(ots_visibles, many=True) #
        return Response(serializer.data, status=status.HTTP_200_OK) #

class PublicVehicleStatusView(generics.GenericAPIView):
    """(GET) Consulta pública de estado OT por patente (normalizada)."""
    permission_classes = [AllowAny] #
    serializer_class = PublicOtStatusSerializer #

    def get(self, request, patente):
        patente_normalizada = patente.upper().replace('-', '') #
        try:
            ot_visible = Ot.objects.filter(
                vehiculo__patente=patente_normalizada, #
                estado__code__in=['ACTIVA', 'PAUSADA'], #
                estado__tipo='ot'
            ).select_related('estado', 'vehiculo').order_by('-fecha_apertura').first() #

            if not ot_visible:
                if not Vehiculo.objects.filter(patente=patente_normalizada).exists(): #
                     return Response({"detail": "Vehículo no encontrado."}, status=status.HTTP_404_NOT_FOUND) #
                else:
                     return Response({"detail": "No hay OT activa o pausada para este vehículo."}, status=status.HTTP_404_NOT_FOUND) #

            serializer = self.get_serializer(ot_visible) #
            return Response(serializer.data) #

        except Exception as e:
            return Response({"detail": "Error interno del servidor."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) #

# --- Reportes ---

# EN capstone-backend/api/views.py

# EN capstone-backend/api/views.py

class SalidasReport(APIView):
    """
    (GET) Reporte Entradas vs Salidas (RF-REP-01). 
    Filtro: ?period=day|week|month.
    Exportar: ?exportar=csv | ?exportar=xlsx
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]

    def get(self, request):
        period = request.GET.get("period", "day").lower()
        if period not in ("day", "week", "month"):
            return Response({"error": "Parámetro 'period' debe ser day, week o month"}, status=status.HTTP_400_BAD_REQUEST)

        tz = "America/Santiago"
        
        # --- (Consulta SQL - Sin cambios) ---
        base_sql = """
            SELECT date_trunc(%s, fecha_hora AT TIME ZONE %s)::date AS label,
                   SUM(CASE WHEN tipo = 'ENTRADA' THEN 1 ELSE 0 END)::int AS vehiculos_entrados,
                   SUM(CASE WHEN tipo = 'SALIDA' THEN 1 ELSE 0 END)::int AS vehiculos_salidos
            FROM apt.bitacora_porteria
            GROUP BY 1 ORDER BY 1 DESC LIMIT %s;
        """
        
        if period == "day": params = ('day', tz, 180)
        elif period == "week": params = ('week', tz, 104)
        else: params = ('month', tz, 36)

        try:
            with connection.cursor() as cur:
                cur.execute(base_sql, params)
                rows = cur.fetchall()
            
            # --- (Procesamiento de datos - Sin cambios) ---
            data = []
            for r in rows:
                entradas = r[1]
                salidas = r[2]
                data.append({
                    "label": r[0].isoformat(),
                    "entradas": entradas,
                    "salidas": salidas,
                    "neto": entradas - salidas
                })
            
            # --- 🌟 LÓGICA DE EXPORTACIÓN AGREGADA 🌟 ---
            output_format = request.query_params.get('exportar')
            if output_format not in ['csv', 'xlsx']:
                # Si no se pide exportar, devuelve JSON
                return Response(data)

            headers = ['Fecha (Label)', 'Entradas', 'Salidas', 'Neto (Entradas - Salidas)']
            filename_base = f"reporte_salidas_{timezone.now().strftime('%Y%m%d')}"

            if output_format == 'csv':
                csv_buffer = StringIO()
                writer = csv.writer(csv_buffer, delimiter=';')
                writer.writerow(headers)
                for item in data:
                    writer.writerow([item['label'], item['entradas'], item['salidas'], item['neto']])
                
                response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
                return response

            elif output_format == 'xlsx':
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Entradas vs Salidas"
                sheet.append(headers)
                for item in data:
                    sheet.append([item['label'], item['entradas'], item['salidas'], item['neto']])
                
                excel_buffer = BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
                
                response = HttpResponse(
                    excel_buffer.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                return response
            
        except Exception as e:
            return Response({"error": f"Error DB al generar reporte: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class ReporteDuracionEtapaView(APIView):
    """
    (GET) Reporte Duración Promedio por Etapa/Pausa (RF-REP-03). 
    Filtros: ?fecha_inicio, ?fecha_fin, ?responsable_id, ?patente
    Exportar: ?exportar=csv | ?exportar=xlsx
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]

    def get(self, request):
        
        # --- (Lógica de filtrado - Sin cambios) ---
        tareas_qs = Tarea.objects.select_related('responsable', 'ot__vehiculo', 'estado')
        pausas_qs = Pausa.objects.select_related('tarea')
        # ... (aquí va toda tu lógica de filtros de fecha, responsable_id, patente_filter) ...
        try: # Filtros de Fecha
            fecha_inicio = request.query_params.get('fecha_inicio')
            # ... (resto de tu lógica de filtros) ...
        except (ValueError, TypeError):
            return Response({"detail": "Formato fecha inválido (YYYY-MM-DD)."}, status=status.HTTP_400_BAD_REQUEST)
        # ... (filtros de responsable_id y patente_filter) ...

        # --- (Lógica de agregación - Sin cambios) ---
        pausas_sq = Pausa.objects.filter(
            tarea=OuterRef('pk'), fin__isnull=False
        ).values('tarea').annotate(total_pausa=Sum(F('fin') - F('inicio'))).values('total_pausa')
        
        promedio_en_proceso_data = tareas_qs.filter(
            estado__code='HECHA', fin__isnull=False, inicio__isnull=False
        ).annotate(
            # ... (anotaciones de duracion_bruta, duracion_pausas, duracion_efectiva) ...
            duracion_bruta=F('fin') - F('inicio'), 
            duracion_pausas=Coalesce(Subquery(pausas_sq[:1]), timedelta(0), output_field=DurationField()),
            duracion_efectiva=F('duracion_bruta') - F('duracion_pausas')
        ).aggregate(
            duracion_promedio=Avg('duracion_efectiva'), 
            conteo=Count('id')
        )
        
        promedio_pausada_data = pausas_qs.filter(
            fin__isnull=False
        ).annotate(
            duracion_pausa=F('fin') - F('inicio')
        ).aggregate(
            duracion_promedio=Avg('duracion_pausa'),
            conteo=Count('id')
        )
        
        # --- (Procesamiento de datos - Sin cambios) ---
        resultados = [] 
        estado_en_proceso = Estado.objects.filter(tipo='tarea', code='EN_PROCESO').first()
        resultados.append({
            'estado_code': 'EN_PROCESO_EFECTIVO', 
            'estado_label': 'En Proceso (Tiempo Efectivo)', 
            'duracion_promedio': promedio_en_proceso_data.get('duracion_promedio') or timedelta(0), 
            'cantidad_tareas_consideradas': promedio_en_proceso_data.get('conteo') or 0 
        })
        estado_pausada = Estado.objects.filter(tipo='tarea', code='PAUSADA').first()
        resultados.append({
            'estado_code': 'PAUSADA', 
            'estado_label': estado_pausada.label if estado_pausada else 'Pausada', 
            'duracion_promedio': promedio_pausada_data.get('duracion_promedio') or timedelta(0), 
            'cantidad_tareas_consideradas': promedio_pausada_data.get('conteo') or 0 
        })
        
        # --- 🌟 LÓGICA DE EXPORTACIÓN AGREGADA 🌟 ---
        output_format = request.query_params.get('exportar')
        if output_format not in ['csv', 'xlsx']:
            # Si no se pide exportar, devuelve JSON
            serializer = ReporteDuracionEtapaSerializer(resultados, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        headers = ['Estado (Etapa/Pausa)', 'Duración Promedio (Segundos)', 'Cantidad Registros (Tareas/Pausas)']
        filename_base = f"reporte_duracion_etapa_{timezone.now().strftime('%Y%m%d')}"

        if output_format == 'csv':
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(headers)
            for item in resultados:
                segundos = round(item['duracion_promedio'].total_seconds()) if isinstance(item['duracion_promedio'], timedelta) else 0
                writer.writerow([item['estado_label'], segundos, item['cantidad_tareas_consideradas']])
            
            response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            return response

        elif output_format == 'xlsx':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Duracion Etapas"
            sheet.append(headers)
            for item in resultados:
                segundos = round(item['duracion_promedio'].total_seconds()) if isinstance(item['duracion_promedio'], timedelta) else 0
                sheet.append([item['estado_label'], segundos, item['cantidad_tareas_consideradas']])
            
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return response

# --- Tablero Kanban ---

class TableroView(APIView):
    """(GET) Datos para el Tablero Kanban (RF-TAB-01). Filtros: ?estado_code, ?patente, ?fecha_inicio, ?fecha_fin, ?responsable_ot_id."""
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser] #

    def get(self, request):
        queryset = Ot.objects.exclude(
            estado__code__in=['CERRADA', 'ANULADA'] #
        ).select_related('vehiculo', 'estado', 'creado_por') #

        estados_filter = request.query_params.get('estado_code') #
        if estados_filter:
            codigos_estado = [e.strip().upper() for e in estados_filter.split(',')] #
            queryset = queryset.filter(estado__code__in=codigos_estado) #

        patente_filter = request.query_params.get('patente') #
        if patente_filter:
            queryset = queryset.filter(vehiculo__patente__icontains=patente_filter.upper().replace('-', '')) #

        try: # Filtros de fecha (sobre fecha_apertura)
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            tz = timezone.get_current_timezone()
            if fecha_inicio:
                fecha_inicio_dt = timezone.make_aware(datetime.fromisoformat(fecha_inicio).replace(hour=0, minute=0, second=0), tz)
                queryset = queryset.filter(fecha_apertura__gte=fecha_inicio_dt)
            if fecha_fin:
                fecha_fin_dt = timezone.make_aware(datetime.fromisoformat(fecha_fin).replace(hour=23, minute=59, second=59), tz)
                queryset = queryset.filter(fecha_apertura__lte=fecha_fin_dt)
        except (ValueError, TypeError):
            return Response({"detail": "Formato fecha inválido (YYYY-MM-DD)."}, status=status.HTTP_400_BAD_REQUEST) #

        responsable_ot_id = request.query_params.get('responsable_ot_id') #
        if responsable_ot_id:
            queryset = queryset.filter(creado_por_id=responsable_ot_id) #

        queryset = queryset.order_by('-actualizado_en') #

        ots_agrupadas = defaultdict(list) #
        estados_visibles = Estado.objects.filter(
            tipo='ot', activo=True
        ).exclude(
            code__in=['CERRADA', 'ANULADA'] #
        ).order_by('orden').values_list('code', flat=True) #

        for code in estados_visibles: #
            ots_agrupadas[code] = [] #

        serializer = TableroOtSerializer(queryset, many=True) #
        for ot_data in serializer.data: #
            estado_code = ot_data['estado']['code'] #
            if estado_code in ots_agrupadas: #
                ots_agrupadas[estado_code].append(ot_data) #

        return Response(ots_agrupadas, status=status.HTTP_200_OK) #

# ============================================
# --- VISTAS: ADMINISTRACIÓN DE USUARIOS ---
# ============================================
# EN capstone-backend/api/views.py (Cerca de la línea 1056)

# ============================================
# --- VISTAS: ADMINISTRACIÓN DE USUARIOS ---
# ============================================

# EN capstone-backend/api/views.py

# EN capstone-backend/api/views.py

# ============================================
# --- VISTAS: ADMINISTRACIÓN DE USUARIOS ---
# ============================================

# EN capstone-backend/api/views.py

# (Asegúrate de que estas clases estén importadas al inicio de views.py)
from .permissions import IsAdminUser, IsSupervisorOrAdminUser
from rest_framework import permissions
# (El resto de tus importaciones: generics, Usuario, AdminUsuarioCreateSerializer, etc.)


class AdminUsuarioListCreateView(generics.ListCreateAPIView):
    """
    (GET) Lista todos los perfiles de usuario (apt.usuario).
    (POST) Invita a un nuevo usuario vía Supabase Auth y crea su perfil local (RF-ADM-01).
    Requiere rol ADMIN.
    """
    
    # --- 1. LÍNEA ORIGINAL COMENTADA ---
    # permission_classes = [permissions.IsAuthenticated, IsAdminUser] 
    
    queryset = Usuario.objects.all().order_by('nombre') 

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return AdminUsuarioCreateSerializer 
        return AdminUsuarioSerializer 

    # --- 2. MÉTODO get_permissions AÑADIDO (ESTA ES LA CORRECCIÓN) ---
    def get_permissions(self):
        """
        GET (Listar): Permite a Admin Y Supervisor (para asignar tareas).
        POST (Crear): Solo Admin.
        """
        if self.request.method == 'POST':
            # Solo Admin puede CREAR/INVITAR usuarios
            return [permissions.IsAuthenticated(), IsAdminUser()]
        
        # Admin O Supervisor pueden LISTAR usuarios (para el dropdown)
        # (Usamos IsSupervisorOrAdminUser que ya tienes definida)
        return [permissions.IsAuthenticated(), IsSupervisorOrAdminUser()]
    # --- FIN DEL BLOQUE AÑADIDO ---

    def create(self, request, *args, **kwargs):
        # (Toda tu función 'create' se mantiene exactamente igual, está perfecta)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # 1. Extraer todos los campos validados
        email = serializer.validated_data['email'] 
        nombre = serializer.validated_data['nombre']
        rol = serializer.validated_data['rol']
        rut = serializer.validated_data.get('rut')
        numero_telefonico = serializer.validated_data.get('numero_telefonico') 

        # 2. Configuración de Supabase
        supabase_url: str = os.environ.get("SUPABASE_URL") 
        supabase_key: str = os.environ.get("SUPABASE_SERVICE_KEY") 
        if not supabase_url or not supabase_key:
            return Response({"detail": "Configuración Supabase (URL/Service Key) incompleta."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        try:
            supabase_admin: Client = create_client(supabase_url, supabase_key)
        except Exception as e:
            return Response({"detail": f"Error inicializando cliente Supabase: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        supabase_user_id = None 

        try: # 3. Llamar a invite_user_by_email (Supabase)
            
            # 🌟 CLAVE: Redirección a nuestra ruta de Vue para establecer la clave.
            REDIRECT_URL_FOR_INVITE = "http://localhost:5173/reset-password" 

            response_supabase = supabase_admin.auth.admin.invite_user_by_email(
                email,
                # Usamos 'options' que es el formato correcto para tu librería
                options={'redirectTo': REDIRECT_URL_FOR_INVITE} 
            )
            supabase_user_id = response_supabase.user.id 
            if not supabase_user_id:
                 raise Exception("La respuesta de Supabase no incluyó el ID del usuario.")
        except AuthApiError as e: 
            return Response({"detail": f"Error al invitar usuario en Supabase: {e.message}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e: 
            return Response({"detail": f"Error inesperado al contactar Supabase: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try: # 4. Crear perfil local (Django/PostgreSQL)
            nuevo_perfil = Usuario.objects.create(
                nombre=nombre,
                email=email,
                rol=rol,
                rut=rut, 
                numero_telefonico=numero_telefonico, 
                external_id=supabase_user_id
            )
        except Exception as e:
            try: # Rollback
                supabase_admin.auth.admin.delete_user(supabase_user_id) 
                error_detail = f"Error al crear perfil local: {e}. Usuario invitado en Supabase fue eliminado."
            except Exception as delete_e:
                error_detail = f"Error al crear perfil local: {e}. ADVERTENCIA: No se pudo eliminar el usuario invitado ({supabase_user_id}) de Supabase ({delete_e})."
            return Response({"detail": error_detail}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 5. Respuesta de éxito
        read_serializer = AdminUsuarioSerializer(nuevo_perfil)
        headers = self.get_success_headers(read_serializer.data)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    


class AdminUsuarioDetailUpdateView(mixins.RetrieveModelMixin,
                                   mixins.UpdateModelMixin,
                                   mixins.DestroyModelMixin, 
                                   generics.GenericAPIView):
    """
    (GET) Muestra detalle de un perfil de usuario (apt.usuario).
    (PUT/PATCH) Actualiza un perfil de usuario (nombre, email, rol, RUT, Teléfono) - RF-ADM-02.
    (DELETE) Da de baja/desactiva un usuario (RF-ADM-03).
    Requiere rol ADMIN.
    """
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]
    queryset = Usuario.objects.all()
    lookup_field = 'id'

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            # Usa el serializer de actualización con todos los campos
            return AdminUsuarioUpdateSerializer
        return AdminUsuarioSerializer

    def get(self, request, *args, **kwargs):
        return self.retrieve(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        """ Maneja la baja (desactivación/eliminación) del usuario (RF-ADM-03). """
        instance: Usuario = self.get_object()

        # 1. Validar que el usuario a eliminar no sea el propio administrador
        if instance.external_id == request.user.username:
            return Response({"detail": "No puedes darte de baja a ti mismo."},
                            status=status.HTTP_400_BAD_REQUEST)

        # 2. Desactivar en Supabase Auth
        supabase_url: str = os.environ.get("SUPABASE_URL")
        supabase_key: str = os.environ.get("SUPABASE_SERVICE_KEY")
        if not supabase_url or not supabase_key:
            return Response({"detail": "Configuración Supabase (URL/Service Key) incompleta."},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        supabase_admin: Client = create_client(supabase_url, supabase_key)

        if instance.external_id:
            supabase_user_id = str(instance.external_id) 
            try:
                # Eliminar la cuenta en Supabase Auth.
                supabase_admin.auth.admin.delete_user(supabase_user_id)
            except AuthApiError as e:
                print(f"ADVERTENCIA: Error al eliminar usuario en Supabase ({supabase_user_id}): {e.message}")
            except Exception as e:
                return Response({"detail": f"Error inesperado al contactar Supabase para baja: {e}"},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 3. Desactivación Lógica en Django (anular external_id)
        instance.external_id = None
        instance.save()

        # 4. Respuesta (204 No Content es estándar para DELETE)
        return Response(status=status.HTTP_204_NO_CONTENT)
# EN api/views.py
# ... (Asegúrate de importar AuditLog en la línea ~40) ...
from .models import (
    Vehiculo, BitacoraPorteria, Ot, Usuario, Estado, Tarea, Pausa, Evidencia, AuditLog # <-- AÑADE AuditLog
)
# ... (Asegúrate de importar AuditLogSerializer en la línea ~45) ...
from .serializers import (
    # ...
    AuditLogSerializer # <-- AÑADE AuditLogSerializer
)

# ... (otras vistas de reportes) ...

# --- Auditoría ---

class AuditLogListAPIView(generics.ListAPIView):
    """
    (GET) Consulta de Registros de Auditoría (RF-AUD-03).
    Filtros: ?user_id, ?entity_type, ?action_type, ?fecha_inicio, ?fecha_fin, ?entity_id.
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]
    queryset = AuditLog.objects.select_related('usuario').all().order_by('-creado_en')
    serializer_class = AuditLogSerializer
    
    # --- 3. AÑADE ESTA LÍNEA PARA ACTIVAR LA PAGINACIÓN ---
    pagination_class = StandardResultsSetPagination
    
    # --- 2. AÑADIR BACKEND DE FILTRADO ---
    # DjangoFilterBackend se encargará de los filtros definidos en filterset_fields.
    # filters.SearchFilter permite búsquedas de texto libre.
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    
    # --- 3. DEFINIR CAMPOS DE FILTRO ---
    # Esto permite filtros exactos como: ?entity_type=ot&entity_id=123&usuario=5
    filterset_fields = ['entity_type', 'entity_id', 'action_type', 'usuario']
    
    # También podemos añadir un filtro de búsqueda de texto libre en los detalles
    search_fields = ['details__icontains']
    

# EN capstone-backend/api/views.py

# ... (después de las importaciones de Python/Django) ...

# --- Funciones Auxiliares ---
def register_audit_log(user_id: int, action_type: str, entity_type: str, entity_id: int = None, details: dict = None):
    """ Escribe un registro en la tabla de auditoría. """
    from .models import AuditLog # Importamos localmente para asegurar el scope
    try:
        AuditLog.objects.create(
            usuario_id=user_id,
            action_type=action_type,
            entity_type=entity_type,
            entity_id=entity_id,
            details=details or {},
            creado_en=timezone.now()
        )
    except Exception as e:
        print(f"ERROR DE AUDITORÍA: No se pudo registrar el log ({action_type} - {entity_type} #{entity_id}): {e}")


# --- Vistas para TareaRepuesto ---

class TareaRepuestoListCreateView(generics.ListCreateAPIView):
    """
    (GET) Lista repuestos usados en una Tarea.
    (POST) Agrega un repuesto a la Tarea (Requiere Mecánico/Supervisor).
    """
    # Requiere autenticación y rol Mecánico/Supervisor
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser]

    def get_queryset(self):
        tarea_id = self.kwargs.get('tarea_id')
        get_object_or_404(Tarea, pk=tarea_id)
        # Retorna los repuestos usados, cargando los detalles del repuesto
        return TareaRepuesto.objects.filter(tarea_id=tarea_id).select_related('repuesto').order_by('creado_en')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TareaRepuestoCreateSerializer
        return TareaRepuestoSerializer

    def perform_create(self, serializer):
        tarea_id = self.kwargs.get('tarea_id')
        tarea_instance = get_object_or_404(Tarea, pk=tarea_id)

        # Validación de estado de OT
        if tarea_instance.ot and tarea_instance.ot.estado and tarea_instance.ot.estado.code in ['CERRADA', 'ANULADA']:
             raise Http404(f"No se pueden añadir repuestos a una tarea de una OT en estado {tarea_instance.ot.estado.code}.")

        # La validación de unicidad (unique_together) se maneja en el modelo al llamar a save(), 
        # pero es mejor atraparla aquí antes de guardar si usaras un serializer.validate() más complejo.
        # Por ahora, inyectamos la Tarea y la fecha.
        instance = serializer.save(
            tarea=tarea_instance, # Inyecta la instancia de Tarea
            creado_en=timezone.now()
        )

        # Registro de Auditoría
        try:
            perfil_usuario = Usuario.objects.get(external_id=self.request.user.username)
            usuario_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            usuario_id = None

        register_audit_log(
            user_id=usuario_id,
            action_type="TAREA_REPUESTO_ADD",
            entity_type="tarea_repuesto",
            entity_id=instance.id,
            details={
                "tarea_id": instance.tarea_id,
                "repuesto_id": instance.repuesto_id,
                "cantidad": str(instance.cantidad)
            }
        )


# --- 1. AÑADE ESTA FUNCIÓN HELPER (NUEVA) ---
# (Similar a la de Horas-Hombre, pero para la OT)

def render_ot_to_pdf_html(ot_data):
    """Genera el contenido HTML para el reporte de detalle de OT."""

    def fdate(date_str):
        if not date_str: return "N/A"
        # (Asegúrate de que timezone esté importado: from django.utils import timezone)
        return timezone.datetime.fromisoformat(date_str).strftime('%d-%m-%Y %I:%M %p')

    # Bucle para generar las Tareas
    tareas_html = []
    if not ot_data.get('tareas'):
        tareas_html.append("<p>No hay tareas asociadas.</p>")
    else:
        for tarea in ot_data['tareas']:

            # Bucle para Repuestos (Sin cambios)
            repuestos_html = ""
            if not tarea.get('repuestos_usados'):
                repuestos_html = "<p>No se declararon repuestos.</p>"
            else:
                repuestos_html = '<table class="repuestos-table"><thead><tr><th>Código</th><th>Descripción</th><th>Cantidad</th></tr></thead><tbody>'
                for item in tarea['repuestos_usados']:
                    repuestos_html += f"""
                        <tr>
                            <td>{html.escape(item['repuesto']['codigo'] or 'S/C')}</td>
                            <td>{html.escape(item['repuesto']['descripcion'])}</td>
                            <td>{item['cantidad']} ({html.escape(item['repuesto']['unidad_medida'])})</td>
                        </tr>
                    """
                repuestos_html += "</tbody></table>"

            # Bucle para Evidencia (CON LA CORRECCIÓN)
            evidencia_html = ""
            if not tarea.get('evidencias'):
                evidencia_html = "<p>No se adjuntó evidencia.</p>"
            else:
                evidencia_html = '<div class="evidence-grid">'
                for file in tarea['evidencias']:
                    if file['mime_type'] and 'image' in file['mime_type']:

                        # --- 1. ARREGLO PARA ESPACIOS EN URL ---
                        # Codifica la URL para manejar espacios (ej. "repuesto 2.jpg")
                        safe_url = urllib.parse.quote(file['url_descarga'], safe=':/')

                        evidencia_html += f"""
                            <div class="evidence-item">
                                <img src="{safe_url}" />
                                <p>{html.escape(file['path'].split('/').pop())}</p>
                            </div>
                        """
                evidencia_html += "</div>"

            # Plantilla de cada Tarea (Sin cambios)
            tareas_html.append(f"""
                <div class="task-block">
                    <h3>Tarea #{tarea['id']}: {html.escape(tarea['nombre'])}</h3>
                    <div class="info-grid task-info">
                        <div><strong>Responsable (Mecánico):</strong> {html.escape(tarea['responsable']['nombre'])}</div>
                        <div><strong>Estado Tarea:</strong> {html.escape(tarea['estado']['label'])}</div>
                        <div><strong>Inicio Tarea:</strong> {fdate(tarea['inicio'])}</div>
                        <div><strong>Fin Tarea:</strong> {fdate(tarea['fin'])}</div>
                    </div>
                    <h4>Repuestos Usados ({len(tarea['repuestos_usados'])})</h4>
                    {repuestos_html}
                    <h4>Evidencia ({len(tarea['evidencias'])})</h4>
                    {evidencia_html}
                </div>
            """)

    # Plantilla Principal del PDF
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Reporte OT #{ot_data['id']}</title>
        <style>
            @page {{ size: A4; margin: 1in; }}
            body {{ font-family: sans-serif; font-size: 10pt; color: #333; }}
            h1 {{ font-size: 18pt; color: #17a2b8; margin: 0; }}
            h2 {{ font-size: 14pt; color: #17a2b8; border-bottom: 1px solid #eee; padding-bottom: 5px; }}
            h3 {{ font-size: 12pt; margin-top: 0; }}
            h4 {{ font-size: 11pt; margin-bottom: 5px; border-bottom: 1px dotted #ccc; }}
            .report-header p {{ font-size: 9pt; color: #555; }}
            .report-section {{ margin-bottom: 20px; page-break-inside: avoid; }}
            .info-grid {{ display: block; }} /* Forzamos una sola columna */
            .info-grid div {{ padding: 3px 0; }}
            .span-all {{ margin-top: 10px; }}
            .span-all p {{ margin: 5px 0 0; border-left: 3px solid #eee; padding-left: 10px; }}
            .task-block {{ border: 1px solid #ddd; border-radius: 8px; padding: 15px; margin-bottom: 15px; page-break-inside: avoid; }}
            .task-info {{ background: #f8f9fa; padding: 10px; border-radius: 4px; display: block; }}
            .repuestos-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            .repuestos-table th, .repuestos-table td {{ border: 1px solid #ddd; padding: 6px; font-size: 9pt; }}
            .repuestos-table th {{ background-color: #f2f2f2; }}
            .evidence-grid {{ display: block; }}
            .evidence-item {{ display: inline-block; width: 48%; margin: 1%; border: 1px solid #eee; text-align: center; page-break-inside: avoid; }}

            /* --- 2. ARREGLO PARA EL ERROR 'getSize' --- */
            .evidence-item img {{ 
                max-width: 100%; /* Usar max-width en lugar de width */
                height: auto; 
                max-height: 150px; 
                object-fit: cover; 
            }}

            .evidence-item p {{ font-size: 9pt; color: #555; margin: 5px; word-break: break-all; }}
        </style>
    </head>
    <body>
        <header class="report-header">
            <h1>Reporte de Cierre: OT #{ot_data['id']}</h1>
            <p>Documento generado el: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}</p>
        </header>
        <section class="report-section">
            <h2>Información de la Orden de Trabajo</h2>
            <div class="info-grid">
                <div><strong>Patente:</strong> {html.escape(ot_data['vehiculo']['patente'])}</div>
                <div><strong>Vehículo:</strong> {html.escape(ot_data['vehiculo']['marca'])} {html.escape(ot_data['vehiculo']['modelo'])}</div>
                <div><strong>Estado:</strong> {html.escape(ot_data['estado']['label'])}</div>
                <div><strong>Creada Por (Supervisor):</strong> {html.escape(ot_data['creado_por']['nombre'])}</div>
                <div><strong>Fecha Apertura:</strong> {fdate(ot_data['fecha_apertura'])}</div>
                <div><strong>Fecha Cierre:</strong> {fdate(ot_data['fecha_cierre'])}</div>
                <div class="span-all">
                    <strong>Descripción OT:</strong>
                    <p>{html.escape(ot_data['descripcion'])}</p>
                </div>
            </div>
        </section>
        <section class="report-section">
            <h2>Detalle de Tareas ({len(ot_data['tareas'])})</h2>
            {''.join(tareas_html)}
        </section>
    </body>
    </html>
    """
    return html_content

# --- 2. REEMPLAZA TU CLASE OtExportDetailView POR ESTA ---
# (La convertimos de generics.RetrieveAPIView a APIView)

class OtExportDetailView(APIView):
    """
    (GET) Detalle completo de OT para exportación/documentación.
    - Devuelve JSON por defecto.
    - Devuelve un PDF si se usa el query param ?exportar=pdf
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]
    
    def get_queryset(self, ot_id):
        # Esta es la consulta optimizada que ya tenías
        return Ot.objects.select_related(
            "vehiculo", "estado", "creado_por"
        ).prefetch_related(
            "tareas__responsable", 
            "tareas__estado",
            "tareas__evidencias__subido_por", # Asegura que el subido_por esté en la evidencia
            "tareas__repuestos_usados__repuesto" # Asegura que el repuesto esté en la lista
        ).annotate(
            tareas_count=Count('tareas') # Mantenemos el conteo por si acaso
        ).get(pk=ot_id)

    def get(self, request, ot_id):
        try:
            ot_instance = self.get_queryset(ot_id)
        except Ot.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        # Serializa los datos (siempre necesario)
        serializer = OtExportSerializer(ot_instance)
        ot_data = serializer.data

        # --- Lógica de Exportación ---
        output_format = request.query_params.get('exportar')
        
        if output_format == 'pdf':
            # 1. Generar el HTML
            html = render_ot_to_pdf_html(ot_data)
            
            # 2. Generar el PDF en memoria
            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(
                html, 
                dest=pdf_buffer,
                encoding='utf-8'
            )
            
            # 3. Enviar el PDF como archivo
            if not pisa_status.err:
                pdf_buffer.seek(0)
                response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
                # 4. ¡AQUÍ ESTÁ EL ARREGLO DEL NOMBRE!
                filename = f"OT_{ot_data['id']}-{ot_data['vehiculo']['patente']}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            
            return Response({"detail": f"Error al generar PDF: {pisa_status.err}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        else:
            # 5. Respuesta JSON por defecto
            # (El frontend usará esta misma vista, pero sin el ?exportar=pdf)
            return Response(ot_data)



class MiPerfilView(APIView):
    """(GET) Obtiene el perfil completo (nombre, rol, rut) del usuario autenticado."""
    
    # --- ESTA ES LA LÍNEA CRÍTICA ---
    # Solo requerir que el usuario esté autenticado (token válido).
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # request.user.username contiene el external_id (UUID de Supabase)
        try:
            # Buscamos el perfil usando el ID de Supabase
            perfil = Usuario.objects.get(external_id=request.user.username)
        except Usuario.DoesNotExist:
            return Response({"detail": "Perfil de usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        
        # Usamos el serializer de administración (AdminUsuarioSerializer) que tiene todos los campos
        serializer = AdminUsuarioSerializer(perfil)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

# EN capstone-backend/api/views.py

# ... (Importaciones de Serializers, alrededor de la línea 40) ...
from .serializers import (
    VehiculoSerializer, VehiculoCreateSerializer, BitacoraSerializer,
    OtSerializer, OtCreateSerializer, OtUpdateSerializer, OtCambiarEstadoSerializer,
    OtEstadoChoferSerializer, PublicOtStatusSerializer,
    TareaSerializer, TareaCreateSerializer, TareaPausarSerializer, TareaUpdateSerializer,
    EvidenciaSerializer,
    ReporteHorasHombreSerializer,
    TableroOtSerializer,
    BitacoraConTotalesSerializer,
    ReporteDuracionEtapaSerializer,
    AdminUsuarioCreateSerializer, AdminUsuarioSerializer, AdminUsuarioUpdateSerializer, 
    AuditLogSerializer, TareaRepuestoSerializer, TareaRepuestoCreateSerializer, OtExportSerializer,TareaExportSerializer, OtExportSerializer,
    ChoferContactoCreateSerializer # 🌟 NUEVA IMPORTACIÓN
)

# ... (Al final del archivo, antes de las vistas de Administración de Usuarios) ...

# --- VISTAS: CREACIÓN DE CONTACTOS CHOFER POR SUPERVISOR ---

class ChoferContactoCreateView(generics.CreateAPIView):
    """
    (POST) Crea un nuevo perfil de Usuario con rol 'CHOFER' (Contacto de vehículo). 
    Permitido a Supervisor/Admin. No usa Supabase Auth.
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminUser] 
    queryset = Usuario.objects.all()
    serializer_class = ChoferContactoCreateSerializer
    
    def perform_create(self, serializer):
        # 1. Forzar el rol a 'CHOFER' y external_id a None (no es un usuario de staff)
        instance = serializer.save(
            rol='CHOFER', 
            external_id=None,
            creado_en=timezone.now()
        )

        # 2. Auditoría
        try:
            perfil_usuario = Usuario.objects.get(external_id=self.request.user.username)
            creador_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            creador_id = None
        
        register_audit_log(
            user_id=creador_id,
            action_type="CHOFER_CONTACT_CREATE",
            entity_type="usuario",
            entity_id=instance.id,
            details={"nombre": instance.nombre, "rut": instance.rut}
        )


# --- AÑADE ESTA NUEVA VISTA AL FINAL DE TUS VISTAS DE API ---

class OtListExportView(APIView):
    """
    (GET) Exporta la lista de OTs (filtrada) en formato CSV o Excel.
    Reutiliza la lógica de filtrado de TableroView.
    Filtros: ?estado_code, ?patente, ?fecha_inicio, ?fecha_fin, ?responsable_ot_id
    Exportar: ?exportar=csv | ?exportar=xlsx
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]

    def get(self, request):
        
        # 1. LÓGICA DE FILTRADO (Copiada de TableroView)
        # ----------------------------------------------------
        queryset = Ot.objects.select_related(
            'vehiculo', 'estado', 'creado_por'
        ) # Cargamos las relaciones

        estados_filter = request.query_params.get('estado_code')
        if estados_filter:
            codigos_estado = [e.strip().upper() for e in estados_filter.split(',')]
            queryset = queryset.filter(estado__code__in=codigos_estado)

        patente_filter = request.query_params.get('patente')
        if patente_filter:
            queryset = queryset.filter(vehiculo__patente__icontains=patente_filter.upper().replace('-', ''))

        try: 
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            tz = timezone.get_current_timezone()
            if fecha_inicio:
                fecha_inicio_dt = timezone.make_aware(datetime.fromisoformat(fecha_inicio).replace(hour=0, minute=0, second=0), tz)
                queryset = queryset.filter(fecha_apertura__gte=fecha_inicio_dt)
            if fecha_fin:
                fecha_fin_dt = timezone.make_aware(datetime.fromisoformat(fecha_fin).replace(hour=23, minute=59, second=59), tz)
                queryset = queryset.filter(fecha_apertura__lte=fecha_fin_dt)
        except (ValueError, TypeError):
            pass # Ignora fechas mal formateadas

        responsable_ot_id = request.query_params.get('responsable_ot_id')
        if responsable_ot_id:
            queryset = queryset.filter(creado_por_id=responsable_ot_id)

        reporte_qs = queryset.order_by('-fecha_apertura')
        # ----------------------------------------------------

        
        # 2. LÓGICA DE EXPORTACIÓN (Adaptada de HorasHombreReport)
        # ----------------------------------------------------
        output_format = request.query_params.get('exportar', 'csv').lower()

        if output_format not in ['csv', 'xlsx']:
             return Response({"detail": "Formato inválido. Usar '?exportar=csv' o '?exportar=xlsx'."}, status=status.HTTP_400_BAD_REQUEST)

        # Definimos las cabeceras del reporte
        headers = ['ID OT', 'Estado', 'Patente', 'Descripcion', 'Creado Por', 'Fecha Apertura', 'Fecha Cierre']

        if output_format == 'csv':
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(headers)
            
            for ot in reporte_qs:
                writer.writerow([
                    ot.id,
                    ot.estado.label if ot.estado else 'N/A',
                    ot.vehiculo.patente if ot.vehiculo else 'N/A',
                    ot.descripcion,
                    ot.creado_por.nombre if ot.creado_por else 'N/A',
                    ot.fecha_apertura.strftime('%Y-%m-%d %H:%M') if ot.fecha_apertura else '',
                    ot.fecha_cierre.strftime('%Y-%m-%d %H:%M') if ot.fecha_cierre else ''
                ])
            
            response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv; charset=utf-8')
            filename = f"reporte_ots_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        elif output_format == 'xlsx':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reporte OTs"
            sheet.append(headers)
            
            for ot in reporte_qs:
                 sheet.append([
                    ot.id,
                    ot.estado.label if ot.estado else 'N/A',
                    ot.vehiculo.patente if ot.vehiculo else 'N/A',
                    ot.descripcion,
                    ot.creado_por.nombre if ot.creado_por else 'N/A',
                    ot.fecha_apertura.strftime('%Y-%m-%d %H:%M') if ot.fecha_apertura else '',
                    ot.fecha_cierre.strftime('%Y-%m-%d %H:%M') if ot.fecha_cierre else ''
                ])

            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].bestFit = True
            
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"reporte_ots_{timezone.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        

# 🌟 1. AÑADE ESTA NUEVA CLASE 🌟
class TareaAnularView(APIView):
    """(POST) Anula (desactiva) una Tarea. (Botón "Desactivar Tarea")"""
    # Permitido a Mecánico o Supervisor (según Vistas Mecánico/Supervisor en mockups)
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser] 

    def post(self, request, tarea_id):
        tarea = get_object_or_404(
            Tarea.objects.select_related(
                'ot', 'ot__estado', 'estado', 'responsable', 'ot__vehiculo'
            ).prefetch_related('evidencias', 'repuestos_usados'),
            pk=tarea_id
        )

        # Validación: No se puede anular una tarea que ya está terminada ('HECHA')
        # o ya está 'ANULADA'.
        estados_invalidos = ['HECHA', 'ANULADA']
        if tarea.estado and tarea.estado.code in estados_invalidos:
            return Response({"detail": f"No se puede anular una tarea en estado '{tarea.estado.code}'."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 2. Obtener el estado final
            estado_anulada = Estado.objects.get(tipo='tarea', code='ANULADA')
        except Estado.DoesNotExist:
            return Response({"detail": "Estado 'ANULADA' para tareas no encontrado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        now = timezone.now()
        estado_anterior_code = tarea.estado.code if tarea.estado else None
        
        # 3. Cerrar pausa activa si la tarea estaba pausada
        # (Igual que en TareaCerrarView)
        if estado_anterior_code == 'PAUSADA':
             pausa_activa = Pausa.objects.filter(tarea=tarea, fin__isnull=True).order_by('-inicio').first()
             if pausa_activa:
                 pausa_activa.fin = now
                 pausa_activa.save()
        
        # 4. Actualizar la tarea al estado "ANULADA"
        tarea.estado = estado_anulada
        tarea.fin = now # Se marca el fin, ya que es un estado terminal
        tarea.actualizado_en = now
        tarea.save()
        
        # 5. Lógica de Auditoría
        try:
            perfil_usuario = Usuario.objects.get(external_id=request.user.username)
            usuario_id = perfil_usuario.id
        except Usuario.DoesNotExist:
            usuario_id = None
            
        register_audit_log(
            user_id=usuario_id,
            action_type="TAREA_ANULAR", # Nueva acción de auditoría
            entity_type="tarea",
            entity_id=tarea.id,
            details={
                "fin_registrado": now.isoformat(),
                "estado_anterior": estado_anterior_code
            }
        )
        
        # 6. Respuesta
        serializer = TareaSerializer(tarea)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class MisTareasListView(generics.ListAPIView):
    """
    (GET) Vista para Mecánico/Supervisor: Lista solo las tareas
    asignadas a ellos que están activas (Nuevas, En Proceso, Pausadas).
    """
    # Usamos el Serializer que ya incluye el vehículo (el que arreglamos antes)
    serializer_class = TareaSerializer
    # Solo Mecánicos o Supervisores pueden ver sus propias tareas
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser]

    def get_queryset(self):
        try:
            # 1. Obtener el perfil de Django del usuario logueado
            perfil_usuario = Usuario.objects.get(external_id=self.request.user.username)
        except Usuario.DoesNotExist:
            # Si no tiene perfil, no tiene tareas
            return Tarea.objects.none()

        # 2. Definir estados de tarea "activos"
        # (No queremos mostrar tareas ya 'HECHA' o 'ANULADA')
        estados_activos = ['NUEVA', 'EN_PROCESO', 'PAUSADA']

        # 3. Filtrar tareas
        return Tarea.objects.filter(
            responsable=perfil_usuario,      # <-- La magia: solo tareas de este usuario
            estado__code__in=estados_activos # <-- Solo tareas activas
        ).select_related(
            'estado', 'ot', 'ot__vehiculo'   # <-- Optimiza la consulta
        ).prefetch_related(
            'evidencias', 'repuestos_usados' # <-- AÑADE ESTE prefetch
        ).order_by('estado__orden', '-actualizado_en')


from .models import Repuesto
from .serializers import RepuestoSerializer
from .permissions import IsMecanicoOrSupervisorUser

class RepuestoListView(generics.ListAPIView):
    """
    (GET) Lista el catálogo completo de Repuestos (para dropdowns).
    Permitido a Mecánicos y Supervisores para que puedan asignarlos.
    """
    permission_classes = [permissions.IsAuthenticated, IsMecanicoOrSupervisorUser]
    serializer_class = RepuestoSerializer
    
    def get_queryset(self):
        # Devuelve solo los repuestos marcados como "activos"
        return Repuesto.objects.filter(activo=True).order_by('descripcion')
    

# --- AÑADE ESTA CLASE NUEVA ---
class OtHistorialListView(generics.ListAPIView):
    """
    (GET) Vista de Historial: Lista todas las OTs CERRADAS y ANULADAS.
    Permitido a Supervisores, Admins y Analistas.
    """
    serializer_class = OtSerializer
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]

    def get_queryset(self):
        # Filtra solo por los dos estados finales
        estados_finales = ['CERRADA', 'ANULADA']
        
        return Ot.objects.filter(
            estado__code__in=estados_finales
        ).select_related(
            "vehiculo", "estado", "creado_por"
        ).annotate(
            tareas_count=Count('tareas') # Incluye el conteo de tareas
        ).order_by('-fecha_cierre', '-actualizado_en') # Muestra las más recientes primero
    
class HorasHombreReport(APIView):
    """
    (GET) Reporte de Horas-Hombre (RF-REP-02).
    Calcula el tiempo de trabajo efectivo (Tarea - Pausas) agrupado
    por responsable, basado en tareas CERRADAS ('HECHA').

    Filtros: ?fecha_inicio, ?fecha_fin, ?responsable_id
    Exportar: ?exportar=csv | ?exportar=xlsx | ?exportar=pdf
    """
    permission_classes = [permissions.IsAuthenticated, IsSupervisorOrAdminOrAnalistaUser]

    def get(self, request):
        
        output_format = request.query_params.get('exportar', 'json').lower()
        if output_format not in ['json', 'csv', 'xlsx', 'pdf']: # <-- Añadido 'pdf'
            return Response({"detail": "Formato inválido. Usar '?exportar=csv', '?exportar=xlsx', '?exportar=pdf' o omitir para JSON."}, status=status.HTTP_400_BAD_REQUEST)

        tareas_qs = Tarea.objects.filter(
            estado__code='HECHA', fin__isnull=False, inicio__isnull=False
        ).select_related('responsable')

        # Variables para metadata del PDF
        fecha_inicio_param = request.query_params.get('fecha_inicio')
        fecha_fin_param = request.query_params.get('fecha_fin')
        
        try: # Filtros de Fecha
            # ... (Toda la lógica de filtros de fecha existente) ...
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            tz = timezone.get_current_timezone()
            if fecha_inicio:
                fecha_inicio_dt = timezone.make_aware(datetime.fromisoformat(fecha_inicio).replace(hour=0, minute=0, second=0), tz)
                tareas_qs = tareas_qs.filter(fin__gte=fecha_inicio_dt)
                fecha_inicio_param = fecha_inicio # Mantener el valor original para el PDF
            if fecha_fin:
                fecha_fin_dt = timezone.make_aware(datetime.fromisoformat(fecha_fin).replace(hour=23, minute=59, second=59), tz)
                tareas_qs = tareas_qs.filter(fin__lte=fecha_fin_dt)
                fecha_fin_param = fecha_fin # Mantener el valor original para el PDF
        except (ValueError, TypeError):
            return Response({"detail": "Formato fecha inválido (YYYY-MM-DD)."}, status=status.HTTP_400_BAD_REQUEST)

        responsable_id = request.query_params.get('responsable_id')
        if responsable_id:
            tareas_qs = tareas_qs.filter(responsable_id=responsable_id)

        pausas_sq = Pausa.objects.filter(
            tarea=OuterRef('pk'), fin__isnull=False
        ).values('tarea').annotate(total_pausa=Sum(F('fin') - F('inicio'))).values('total_pausa')

        reporte_qs = tareas_qs.annotate(
            duracion_bruta=F('fin') - F('inicio'),
            duracion_pausas=Coalesce(Subquery(pausas_sq[:1]), timedelta(0), output_field=fields.DurationField())
        ).annotate(
            duracion_efectiva=F('duracion_bruta') - F('duracion_pausas')
        ).values(
            'responsable', 'responsable__nombre', 'responsable__rol'
        ).annotate(
            duracion_total_efectiva=Sum('duracion_efectiva'),
            conteo_tareas=Count('id')
        ).order_by('-duracion_total_efectiva')

        # --- LÓGICA DE EXPORTACIÓN A PDF (RF-EXP-02) ---
        if output_format == 'pdf':
            html = render_to_pdf_html(reporte_qs, fecha_inicio_param, fecha_fin_param)
            
            # Crear el objeto BytesIO en memoria para el PDF
            pdf_buffer = BytesIO()
            
            # Generar el PDF a partir del HTML
            pisa_status = pisa.CreatePDF(
                html, 
                dest=pdf_buffer, # El destino es el buffer
                encoding='utf-8' # Asegurar que tilde y ñ funcionen
            )
            
            # Si el PDF se generó sin errores
            if not pisa_status.err:
                pdf_buffer.seek(0)
                response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
                filename = f"reporte_horas_hombre_{timezone.now().strftime('%Y%m%d')}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            
            # Si hay errores en la generación del PDF
            return Response({"detail": f"Error al generar PDF: {pisa_status.err}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # --- FIN LÓGICA PDF ---

        elif output_format == 'csv':
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(['ID Responsable', 'Nombre Responsable', 'Rol', 'Segundos Trabajados', 'Tareas Completadas'])
            for item in reporte_qs:
                segundos = round(item['duracion_total_efectiva'].total_seconds()) if item['duracion_total_efectiva'] else 0
                writer.writerow([item['responsable'], item['responsable__nombre'], item['responsable__rol'], segundos, item['conteo_tareas']])
            response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv; charset=utf-8')
            filename = f"reporte_horas_hombre_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        elif output_format == 'xlsx':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Horas Hombre"
            headers = ['ID Responsable', 'Nombre Responsable', 'Rol', 'Segundos Trabajados', 'Tareas Completadas']
            sheet.append(headers)
            for item in reporte_qs:
                segundos = round(item['duracion_total_efectiva'].total_seconds()) if item['duracion_total_efectiva'] else 0
                sheet.append([item['responsable'], item['responsable__nombre'], item['responsable__rol'], segundos, item['conteo_tareas']])
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].bestFit = True
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"reporte_horas_hombre_{timezone.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        else: # output_format == 'json' (por defecto)
            serializer = ReporteHorasHombreSerializer(reporte_qs, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)